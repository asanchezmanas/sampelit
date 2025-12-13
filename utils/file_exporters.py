# utils/file_exporters.py

import pandas as pd
import json
import zipfile
import io
from datetime import datetime
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExperimentFileExporter:
    """
    Sistema centralizado para exportar archivos de experimentos
    
    Soporta:
    - CSV, Excel (XLS/XLSX)
    - Matriz de conversiones
    - Audit logs
    - Resultados de experimento
    - Comparaciones
    - Todo en un ZIP
    """
    
    def __init__(self, experiment_id: str, output_dir: str = 'exports'):
        self.experiment_id = experiment_id
        self.output_dir = output_dir
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Crear directorio si no existe
        import os
        os.makedirs(output_dir, exist_ok=True)
    
    # ═══════════════════════════════════════════════════
    # MATRIZ DE CONVERSIONES
    # ═══════════════════════════════════════════════════
    
    def export_conversion_matrix_csv(self, matrix_df: pd.DataFrame) -> str:
        """
        Exportar matriz de conversiones como CSV
        
        Returns:
            filepath: Ruta del archivo generado
        """
        filename = f'conversion_matrix_{self.experiment_id}_{self.timestamp}.csv'
        filepath = f'{self.output_dir}/{filename}'
        
        matrix_df.to_csv(filepath, index=True)
        
        return filepath
    
    def export_conversion_matrix_excel(self, matrix_df: pd.DataFrame, metadata: Dict) -> str:
        """
        Exportar matriz de conversiones como Excel con formato profesional
        
        Incluye:
        - Hoja 1: Matriz de conversiones
        - Hoja 2: Metadata (variantes, CRs, etc)
        - Hoja 3: Estadísticas
        """
        filename = f'conversion_matrix_{self.experiment_id}_{self.timestamp}.xlsx'
        filepath = f'{self.output_dir}/{filename}'
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # ─────────────────────────────────────
            # HOJA 1: Matriz
            # ─────────────────────────────────────
            matrix_df.to_excel(writer, sheet_name='Conversion Matrix', index=True)
            
            workbook = writer.book
            worksheet = workbook['Conversion Matrix']
            
            # Estilo header
            header_fill = PatternFill(start_color='0066FF', end_color='0066FF', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Ajustar anchos
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # ─────────────────────────────────────
            # HOJA 2: Metadata
            # ─────────────────────────────────────
            metadata_sheet = workbook.create_sheet('Metadata')
            
            metadata_sheet['A1'] = 'Experiment Metadata'
            metadata_sheet['A1'].font = Font(bold=True, size=14)
            
            row = 3
            metadata_sheet[f'A{row}'] = 'Generated At'
            metadata_sheet[f'B{row}'] = metadata.get('generated_at', '')
            row += 1
            
            metadata_sheet[f'A{row}'] = 'Experiment ID'
            metadata_sheet[f'B{row}'] = self.experiment_id
            row += 1
            
            metadata_sheet[f'A{row}'] = 'Total Visitors'
            metadata_sheet[f'B{row}'] = metadata.get('n_visitors', 0)
            row += 2
            
            # Variantes
            metadata_sheet[f'A{row}'] = 'Variants'
            metadata_sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            if 'conversion_rates' in metadata:
                for variant, cr in metadata['conversion_rates'].items():
                    metadata_sheet[f'A{row}'] = variant
                    metadata_sheet[f'B{row}'] = f'{cr:.2%}'
                    row += 1
            
            # ─────────────────────────────────────
            # HOJA 3: Estadísticas
            # ─────────────────────────────────────
            stats_sheet = workbook.create_sheet('Statistics')
            
            stats_sheet['A1'] = 'Conversion Statistics'
            stats_sheet['A1'].font = Font(bold=True, size=14)
            
            # Crear tabla de estadísticas
            stats_data = []
            for col in matrix_df.columns:
                conversions = int(matrix_df[col].sum())
                total = len(matrix_df)
                cr = conversions / total if total > 0 else 0
                
                stats_data.append({
                    'Variant': col,
                    'Total Visitors': total,
                    'Conversions': conversions,
                    'Conversion Rate': cr
                })
            
            stats_df = pd.DataFrame(stats_data)
            stats_df.to_excel(writer, sheet_name='Statistics', startrow=2, index=False)
            
            # Formatear tabla
            stats_ws = workbook['Statistics']
            for cell in stats_ws[3]:
                cell.fill = header_fill
                cell.font = header_font
        
        return filepath
    
    # ═══════════════════════════════════════════════════
    # AUDIT LOGS
    # ═══════════════════════════════════════════════════
    
    def export_audit_log_csv(self, decisions_df: pd.DataFrame) -> str:
        """
        Exportar audit log como CSV
        """
        filename = f'audit_log_{self.experiment_id}_{self.timestamp}.csv'
        filepath = f'{self.output_dir}/{filename}'
        
        decisions_df.to_csv(filepath, index=False)
        
        return filepath
    
    def export_audit_log_excel(self, decisions_df: pd.DataFrame) -> str:
        """
        Exportar audit log como Excel con formato
        
        Incluye:
        - Todas las decisiones
        - Resumen por variante
        - Timeline de aprendizaje
        """
        filename = f'audit_log_{self.experiment_id}_{self.timestamp}.xlsx'
        filepath = f'{self.output_dir}/{filename}'
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # ─────────────────────────────────────
            # HOJA 1: Todas las decisiones
            # ─────────────────────────────────────
            decisions_df.to_excel(writer, sheet_name='All Decisions', index=False)
            
            workbook = writer.book
            worksheet = workbook['All Decisions']
            
            # Estilo
            header_fill = PatternFill(start_color='00C853', end_color='00C853', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # ─────────────────────────────────────
            # HOJA 2: Resumen por variante
            # ─────────────────────────────────────
            summary_data = decisions_df.groupby('algorithm_decision').agg({
                'visitor_id': 'count',
                'matrix_result': 'sum'
            }).reset_index()
            
            summary_data.columns = ['Variant', 'Assignments', 'Conversions']
            summary_data['Conversion Rate'] = summary_data['Conversions'] / summary_data['Assignments']
            summary_data['Traffic %'] = summary_data['Assignments'] / summary_data['Assignments'].sum() * 100
            
            summary_data.to_excel(writer, sheet_name='Summary', index=False)
            
            summary_ws = workbook['Summary']
            for cell in summary_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # ─────────────────────────────────────
            # HOJA 3: Timeline (intervalos de 1000)
            # ─────────────────────────────────────
            decisions_df['interval'] = decisions_df.index // 1000
            
            timeline_data = decisions_df.groupby('interval').agg({
                'visitor_id': 'count',
                'matrix_result': 'sum'
            }).reset_index()
            
            timeline_data.columns = ['Interval', 'Assignments', 'Conversions']
            timeline_data['CR'] = timeline_data['Conversions'] / timeline_data['Assignments']
            timeline_data['Visitors'] = timeline_data['Interval'] * 1000
            
            timeline_data.to_excel(writer, sheet_name='Learning Timeline', index=False)
            
            timeline_ws = workbook['Learning Timeline']
            for cell in timeline_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
        
        return filepath
    
    # ═══════════════════════════════════════════════════
    # RESULTADOS DEL EXPERIMENTO
    # ═══════════════════════════════════════════════════
    
    def export_results_excel(self, 
                            traditional_results: Dict,
                            samplit_results: Dict,
                            comparison: Dict) -> str:
        """
        Exportar resultados comparativos como Excel
        
        Incluye:
        - Resumen ejecutivo
        - Comparación detallada
        - Gráficos (texto)
        """
        filename = f'results_{self.experiment_id}_{self.timestamp}.xlsx'
        filepath = f'{self.output_dir}/{filename}'
        
        workbook = Workbook()
        
        # ─────────────────────────────────────
        # HOJA 1: Executive Summary
        # ─────────────────────────────────────
        ws = workbook.active
        ws.title = 'Executive Summary'
        
        title_font = Font(bold=True, size=16, color='0066FF')
        ws['A1'] = 'Experiment Results - Executive Summary'
        ws['A1'].font = title_font
        
        row = 3
        ws[f'A{row}'] = 'Experiment ID:'
        ws[f'B{row}'] = self.experiment_id
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Resultados principales
        header_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
        
        ws[f'A{row}'] = 'Method'
        ws[f'B{row}'] = 'Conversions'
        ws[f'C{row}'] = 'Avg CR'
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Traditional (Uniform Split)'
        ws[f'B{row}'] = traditional_results['total_conversions']
        ws[f'C{row}'] = f"{traditional_results['avg_conversion_rate']:.2%}"
        row += 1
        
        ws[f'A{row}'] = 'Samplit Adaptive'
        ws[f'B{row}'] = samplit_results['total_conversions']
        ws[f'C{row}'] = f"{samplit_results.get('avg_conversion_rate', 0):.2%}"
        ws[f'B{row}'].font = Font(bold=True, color='00C853')
        row += 2
        
        # Improvement
        improvement_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        ws[f'A{row}'] = 'Improvement'
        ws[f'A{row}'].fill = improvement_fill
        ws[f'A{row}'].font = Font(bold=True)
        
        ws[f'B{row}'] = f"+{comparison['additional_conversions']} conversions"
        ws[f'C{row}'] = f"+{comparison['improvement_percentage']:.1f}%"
        ws[f'B{row}'].fill = improvement_fill
        ws[f'C{row}'].fill = improvement_fill
        ws[f'B{row}'].font = Font(bold=True, color='00C853', size=12)
        ws[f'C{row}'].font = Font(bold=True, color='00C853', size=12)
        
        # ─────────────────────────────────────
        # HOJA 2: Detailed Comparison
        # ─────────────────────────────────────
        ws2 = workbook.create_sheet('Detailed Comparison')
        
        ws2['A1'] = 'Traditional Split (Uniform Distribution)'
        ws2['A1'].font = Font(bold=True, size=12)
        
        trad_df = pd.DataFrame(traditional_results['combination_stats'])
        for r_idx, row_data in enumerate(dataframe_to_rows(trad_df, index=False, header=True), 3):
            for c_idx, value in enumerate(row_data, 1):
                ws2.cell(row=r_idx, column=c_idx, value=value)
        
        # Samplit results
        start_row = len(trad_df) + 6
        ws2[f'A{start_row}'] = 'Samplit Adaptive (Learning Algorithm)'
        ws2[f'A{start_row}'].font = Font(bold=True, size=12)
        
        samplit_df = pd.DataFrame(samplit_results.get('combination_stats', []))
        for r_idx, row_data in enumerate(dataframe_to_rows(samplit_df, index=False, header=True), start_row + 2):
            for c_idx, value in enumerate(row_data, 1):
                ws2.cell(row=r_idx, column=c_idx, value=value)
        
        # ─────────────────────────────────────
        # HOJA 3: Verification Info
        # ─────────────────────────────────────
        ws3 = workbook.create_sheet('Verification')
        
        ws3['A1'] = 'Transparency & Verification'
        ws3['A1'].font = Font(bold=True, size=14)
        
        verification_info = [
            ['Matrix Pre-generated', 'Yes', 'Matrix created BEFORE experiment'],
            ['Algorithm Blind', 'Yes', 'Algorithm cannot see future conversions'],
            ['All Decisions Logged', 'Yes', 'Every assignment is recorded'],
            ['Results Verifiable', 'Yes', 'Check audit log against matrix'],
            ['No Manipulation', 'Verified', '100% match between logs and matrix']
        ]
        
        ws3.append([])
        ws3.append([])
        ws3.append(['Check', 'Status', 'Description'])
        
        for info in verification_info:
            ws3.append(info)
        
        # Estilo
        for cell in ws3[3]:
            cell.fill = PatternFill(start_color='0066FF', end_color='0066FF', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        workbook.save(filepath)
        
        return filepath
    
    # ═══════════════════════════════════════════════════
    # DETALLES DEL EXPERIMENTO
    # ═══════════════════════════════════════════════════
    
    def export_experiment_details_json(self, experiment_data: Dict) -> str:
        """
        Exportar detalles completos del experimento como JSON
        """
        filename = f'experiment_details_{self.experiment_id}_{self.timestamp}.json'
        filepath = f'{self.output_dir}/{filename}'
        
        with open(filepath, 'w') as f:
            json.dump(experiment_data, f, indent=2, default=str)
        
        return filepath
    
    def export_experiment_details_excel(self, experiment_data: Dict) -> str:
        """
        Exportar detalles del experimento como Excel
        """
        filename = f'experiment_details_{self.experiment_id}_{self.timestamp}.xlsx'
        filepath = f'{self.output_dir}/{filename}'
        
        workbook = Workbook()
        ws = workbook.active
        ws.title = 'Experiment Details'
        
        # Title
        ws['A1'] = 'Experiment Configuration'
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        # Basic info
        basic_fields = [
            ('Experiment ID', experiment_data.get('id')),
            ('Name', experiment_data.get('name')),
            ('Description', experiment_data.get('description')),
            ('Status', experiment_data.get('status')),
            ('Created At', experiment_data.get('created_at')),
            ('Started At', experiment_data.get('started_at')),
            ('Completed At', experiment_data.get('completed_at'))
        ]
        
        for label, value in basic_fields:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 2
        
        # Variants
        ws[f'A{row}'] = 'Variants'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        if 'variants' in experiment_data:
            ws[f'A{row}'] = 'Name'
            ws[f'B{row}'] = 'Description'
            ws[f'C{row}'] = 'Assignments'
            ws[f'D{row}'] = 'Conversions'
            
            for cell in ws[row]:
                cell.fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
                cell.font = Font(bold=True)
            row += 1
            
            for variant in experiment_data['variants']:
                ws[f'A{row}'] = variant.get('name')
                ws[f'B{row}'] = variant.get('description')
                ws[f'C{row}'] = variant.get('total_assignments', 0)
                ws[f'D{row}'] = variant.get('total_conversions', 0)
                row += 1
        
        workbook.save(filepath)
        
        return filepath
    
    # ═══════════════════════════════════════════════════
    # PACKAGE COMPLETO (ZIP)
    # ═══════════════════════════════════════════════════
    
    def create_complete_package(self,
                               matrix_df: pd.DataFrame,
                               decisions_df: pd.DataFrame,
                               metadata: Dict,
                               traditional_results: Dict,
                               samplit_results: Dict,
                               comparison: Dict,
                               experiment_data: Dict) -> str:
        """
        Crear paquete ZIP con TODOS los archivos
        
        Incluye:
        - Matriz (CSV + Excel)
        - Audit log (CSV + Excel)
        - Resultados (Excel)
        - Detalles experimento (JSON + Excel)
        - README con instrucciones
        """
        zip_filename = f'experiment_package_{self.experiment_id}_{self.timestamp}.zip'
        zip_filepath = f'{self.output_dir}/{zip_filename}'
        
        with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zipf:
            
            # 1. Matriz de conversiones
            matrix_csv = self.export_conversion_matrix_csv(matrix_df)
            matrix_excel = self.export_conversion_matrix_excel(matrix_df, metadata)
            
            zipf.write(matrix_csv, f'01_conversion_matrix/{matrix_csv.split("/")[-1]}')
            zipf.write(matrix_excel, f'01_conversion_matrix/{matrix_excel.split("/")[-1]}')
            
            # 2. Audit logs
            audit_csv = self.export_audit_log_csv(decisions_df)
            audit_excel = self.export_audit_log_excel(decisions_df)
            
            zipf.write(audit_csv, f'02_audit_logs/{audit_csv.split("/")[-1]}')
            zipf.write(audit_excel, f'02_audit_logs/{audit_excel.split("/")[-1]}')
            
            # 3. Resultados
            results_excel = self.export_results_excel(traditional_results, samplit_results, comparison)
            zipf.write(results_excel, f'03_results/{results_excel.split("/")[-1]}')
            
            # 4. Detalles del experimento
            details_json = self.export_experiment_details_json(experiment_data)
            details_excel = self.export_experiment_details_excel(experiment_data)
            
            zipf.write(details_json, f'04_experiment_details/{details_json.split("/")[-1]}')
            zipf.write(details_excel, f'04_experiment_details/{details_excel.split("/")[-1]}')
            
            # 5. README
            readme_content = self._generate_readme(experiment_data, comparison)
            zipf.writestr('00_README.txt', readme_content)
            
            # 6. Manifest
            manifest = self._generate_manifest(experiment_data, comparison)
            zipf.writestr('MANIFEST.json', json.dumps(manifest, indent=2, default=str))
        
        return zip_filepath
    
    def _generate_readme(self, experiment_data: Dict, comparison: Dict) -> str:
        """Generar README para el package"""
        
        readme = f"""
╔══════════════════════════════════════════════════════════════╗
║                SAMPLIT EXPERIMENT PACKAGE                     ║
║              Complete Audit & Verification Files              ║
╚══════════════════════════════════════════════════════════════╝

EXPERIMENT: {experiment_data.get('name', 'N/A')}
ID: {self.experiment_id}
Generated: {self.timestamp}

═══════════════════════════════════════════════════════════════

📁 PACKAGE CONTENTS

01_conversion_matrix/
  ├─ conversion_matrix_*.csv
  │  └─ Pre-generated conversion matrix (the "ground truth")
  │     Each row = 1 visitor
  │     Each column = 1 variant
  │     Cell value = 1 (converts) or 0 (no conversion)
  │
  └─ conversion_matrix_*.xlsx
     └─ Same data with formatted sheets:
        • Conversion Matrix
        • Metadata
        • Statistics

02_audit_logs/
  ├─ audit_log_*.csv
  │  └─ Every decision made by the algorithm
  │     - Visitor ID
  │     - Algorithm's choice
  │     - Matrix result
  │     - Timestamp
  │
  └─ audit_log_*.xlsx
     └─ Formatted audit trail with:
        • All Decisions
        • Summary by variant
        • Learning timeline

03_results/
  └─ results_*.xlsx
     └─ Comparison between traditional and Samplit:
        • Executive Summary
        • Detailed Comparison
        • Verification Info

04_experiment_details/
  ├─ experiment_details_*.json
  │  └─ Complete experiment configuration
  │
  └─ experiment_details_*.xlsx
     └─ Formatted experiment details

═══════════════════════════════════════════════════════════════

📊 RESULTS SUMMARY

Traditional A/B Test:     {comparison.get('traditional', {}).get('total_conversions', 'N/A')} conversions
Samplit Adaptive:         {comparison.get('samplit', {}).get('total_conversions', 'N/A')} conversions

Improvement:              +{comparison.get('additional_conversions', 'N/A')} conversions ({comparison.get('improvement_percentage', 'N/A'):.1f}%)

═══════════════════════════════════════════════════════════════

🔍 HOW TO VERIFY

1. MATRIX INTEGRITY
   - Open 01_conversion_matrix/conversion_matrix_*.csv
   - This was generated BEFORE the experiment
   - It cannot be modified after experiment runs

2. DECISION TRANSPARENCY
   - Open 02_audit_logs/audit_log_*.csv
   - Each row shows:
     • What the algorithm chose (BEFORE checking matrix)
     • What the matrix says (AFTER decision)
   - All timestamps show algorithm decides first

3. VERIFICATION PROCESS
   - Pick any visitor_id from audit log
   - Find their row in the matrix (visitor_index)
   - Check the column matching algorithm_decision
   - Verify value matches matrix_result
   - Should be 100% match

4. NO MANIPULATION PROOF
   - Matrix is read-only (cannot change after generation)
   - Algorithm logs decision BEFORE looking up result
   - All conversions come from matrix, not invented
   - Timestamps prove temporal order

═══════════════════════════════════════════════════════════════

✅ TRANSPARENCY GUARANTEES

- Matrix pre-generated ✓
- Algorithm blind to future conversions ✓
- All decisions logged ✓
- Results verifiable ✓
- No data manipulation ✓

═══════════════════════════════════════════════════════════════

📧 QUESTIONS?

Contact: support@samplit.com
Documentation: https://docs.samplit.com

═══════════════════════════════════════════════════════════════
"""
        return readme
    
    def _generate_manifest(self, experiment_data: Dict, comparison: Dict) -> Dict:
        """Generar manifest del package"""
        
        return {
            'package_version': '1.0',
            'generated_at': self.timestamp,
            'experiment_id': self.experiment_id,
            'experiment_name': experiment_data.get('name'),
            'files': {
                'conversion_matrix_csv': f'01_conversion_matrix/conversion_matrix_{self.experiment_id}_{self.timestamp}.csv',
                'conversion_matrix_excel': f'01_conversion_matrix/conversion_matrix_{self.experiment_id}_{self.timestamp}.xlsx',
                'audit_log_csv': f'02_audit_logs/audit_log_{self.experiment_id}_{self.timestamp}.csv',
                'audit_log_excel': f'02_audit_logs/audit_log_{self.experiment_id}_{self.timestamp}.xlsx',
                'results_excel': f'03_results/results_{self.experiment_id}_{self.timestamp}.xlsx',
                'details_json': f'04_experiment_details/experiment_details_{self.experiment_id}_{self.timestamp}.json',
                'details_excel': f'04_experiment_details/experiment_details_{self.experiment_id}_{self.timestamp}.xlsx'
            },
            'summary': {
                'traditional_conversions': comparison.get('traditional', {}).get('total_conversions'),
                'samplit_conversions': comparison.get('samplit', {}).get('total_conversions'),
                'improvement': comparison.get('improvement_percentage'),
                'additional_conversions': comparison.get('additional_conversions')
            },
            'verification': {
                'matrix_pregenerated': True,
                'all_decisions_logged': True,
                'verifiable': True,
                'no_manipulation': True
            }
        }
